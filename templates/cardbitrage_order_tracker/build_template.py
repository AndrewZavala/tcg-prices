#!/usr/bin/env python3
"""Build a multi-tab Excel file for import into Google Sheets.

Usage:
    pip install openpyxl
    python build_template.py

Output: Cardbitrage_Order_Tracker.xlsx in this directory.
"""

from __future__ import annotations

from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError as exc:
    raise SystemExit("Install openpyxl first: pip install openpyxl") from exc

DIR = Path(__file__).resolve().parent
OUT = DIR / "Cardbitrage_Order_Tracker.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="334155")
HEADER_FONT = Font(bold=True, color="FFFFFF")
FREEZE = "A2"

ORDERS_HEADERS = [
    "tcg_order_id",
    "seller",
    "status",
    "ordered_on",
    "shipped_on",
    "delivered_on",
    "tracking",
    "buy_total",
    "shipping",
    "checkout_key",
    "simplified_lot",
    "ck_batch_id",
    "last_email_at",
    "last_email_subject",
    "notes",
]

ORDERS_EXAMPLE = [
    "2D8EFC37-FBA269-02EAE",
    "Gamer's Empyrean",
    "Ordered",
    "2026-07-02",
    "",
    "",
    "",
    45.48,
    "",
    "2026-07-02:165.79",
    "gamers-empyrean|relentless-assault",
    "",
    "2026-07-02",
    "Your TCGplayer.com order ... has processed.",
    "Example row — delete after import",
]

TRANSACTIONS_HEADERS = [
    "transaction_id",
    "email_date",
    "email_type",
    "source",
    "reference",
    "amount_usd",
    "raw_subject",
    "message_id",
    "notes",
]

TRANSACTIONS_EXAMPLE = [
    "2026-07-02:165.79:processed",
    "2026-07-02",
    "tcg_processed",
    "tcgplayer",
    "2026-07-02:165.79",
    165.79,
    "Your TCGplayer.com order of Relentless Assault...",
    "",
    "Example row — delete after import",
]

CK_HEADERS = [
    "ck_batch_id",
    "label",
    "sent_on",
    "ck_ref",
    "expected_ck_usd",
    "status",
    "notes",
]

CK_EXAMPLE = [
    "CK-2026-07-02-A",
    "Tarkan July lot",
    "",
    "",
    199.34,
    "Draft",
    "Example row — delete after import",
]

CARDS_HEADERS = [
    "tcg_order_id",
    "checkout_key",
    "card_name",
    "set_name",
    "finish",
    "condition",
    "buy_qty",
    "unit_buy_usd",
    "line_total_usd",
    "ck_adj_usd",
    "notes",
]

CARDS_EXAMPLE = [
    ["2D8EFC37-FBA269-02EAE", "2026-07-02:165.79", "Relentless Assault", "7th Edition", "foil", "MP", 1, 45.48, 45.48, "", "Example"],
    ["", "2026-07-02:165.79", "Blasphemous Act (Extended Art)", "Commander Legends", "foil", "NM", 8, "", "", "", "Unassigned until shipped email"],
]

ORDER_STATUS = "Ordered,Shipped,Delivered,Packed,Sent to CK,Done"
CK_STATUS = "Draft,Sent,Received by CK,Paid,Closed"

COL_WIDTHS = {
    "Transactions": [28, 12, 16, 12, 20, 10, 40, 24, 28],
    "Orders": [22, 22, 14, 12, 12, 12, 18, 10, 10, 18, 36, 16, 18, 36, 28],
    "CK_batches": [18, 24, 12, 14, 14, 16, 28],
    "Cards": [22, 18, 22, 20, 10, 10, 8, 10, 10, 10, 24],
}


def style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = FREEZE
    ws.row_dimensions[1].height = 28


def add_validation(ws, col_letter: str, formula: str, max_row: int = 5000) -> None:
    dv = DataValidation(type="list", formula1=f'"{formula}"', allow_blank=True)
    dv.error = "Pick a value from the list"
    dv.errorTitle = "Invalid status"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{max_row}")


def setup_sheet(ws, headers: list[str], example_rows: list, status_col: str | None, status_list: str | None):
    ws.append(headers)
    if example_rows:
        if example_rows and isinstance(example_rows[0], list):
            for row in example_rows:
                ws.append(row)
        else:
            ws.append(example_rows)
    style_header(ws, len(headers))
    widths = COL_WIDTHS.get(ws.title, [14] * len(headers))
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    if status_col and status_list:
        add_validation(ws, status_col, status_list)


def main() -> None:
    wb = Workbook()
    txn = wb.active
    txn.title = "Transactions"
    setup_sheet(txn, TRANSACTIONS_HEADERS, TRANSACTIONS_EXAMPLE, None, None)

    orders = wb.create_sheet("Orders")
    setup_sheet(orders, ORDERS_HEADERS, ORDERS_EXAMPLE, "C", ORDER_STATUS)

    ck = wb.create_sheet("CK_batches")
    setup_sheet(ck, CK_HEADERS, CK_EXAMPLE, "F", CK_STATUS)

    cards = wb.create_sheet("Cards")
    setup_sheet(cards, CARDS_HEADERS, CARDS_EXAMPLE, None, None)

    readme = wb.create_sheet("_README")
    readme.sheet_state = "hidden"
    lines = [
        "Cardbitrage order tracker",
        "",
        "Transactions = one row per email event (audit log)",
        "Orders.tcg_order_id = upsert key for TCG seller orders",
        "Orders.checkout_key = groups multi-seller checkouts from one processed email",
        "Orders.ck_batch_id links to CK_batches.ck_batch_id",
        "Cards.checkout_key links line items before tcg_order_id is known",
        "",
        "Delete example rows on Transactions, Orders, CK_batches, and Cards after import.",
        "See templates/cardbitrage_order_tracker/README.md in the TCG repo.",
    ]
    for i, line in enumerate(lines, start=1):
        readme.cell(row=i, column=1, value=line)

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
